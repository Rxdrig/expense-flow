from io import BytesIO
from decimal import Decimal

from django.utils import timezone
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def _format_currency(value):
    amount = Decimal(value or 0)
    return f"${amount:,.0f}".replace(",", ".")


def _format_date(value):
    if not value:
        return "-"
    return value.strftime("%d/%m/%Y")


def build_pdf_report(user, expenses, generated_at=None):
    generated_at = generated_at or timezone.localdate()
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ExpenseFlowTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=18,
        textColor=colors.HexColor("#1f2a44"),
        spaceAfter=8,
    )
    meta_style = ParagraphStyle(
        "ExpenseFlowMeta",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        textColor=colors.HexColor("#4f5f8b"),
        spaceAfter=12,
    )

    total = sum((expense.amount for expense in expenses), Decimal("0"))
    rows = [["Título", "Categoría", "Fecha", "Descripción", "Monto"]]

    for expense in expenses:
        rows.append([
            expense.title,
            expense.get_category_display(),
            _format_date(expense.date),
            expense.description or "-",
            _format_currency(expense.amount),
        ])

    if len(rows) == 1:
        rows.append(["Sin gastos para el rango seleccionado", "", "", "", ""])

    table = Table(rows, repeatRows=1, colWidths=[70 * mm, 38 * mm, 24 * mm, 90 * mm, 28 * mm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#8fa5ff")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("LEADING", (0, 0), (-1, -1), 11),
                ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f7ff")]),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#c9d4ff")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (4, 1), (4, -1), "RIGHT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )

    elements = [
        Paragraph("ExpenseFlow - Reporte de gastos", title_style),
        Paragraph(f"Usuario: {user.get_full_name() or user.username}", meta_style),
        Paragraph(f"Fecha de generación: {generated_at.strftime('%d/%m/%Y')}", meta_style),
        Spacer(1, 6),
        table,
        Spacer(1, 8),
        Paragraph(f"Total general: {_format_currency(total)}", meta_style),
    ]
    doc.build(elements)
    return buffer.getvalue()


def build_xlsx_report(user, expenses, generated_at=None):
    generated_at = generated_at or timezone.localdate()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Reporte"

    headers = ["Título", "Categoría", "Fecha", "Descripción", "Monto"]
    worksheet.append(["ExpenseFlow - Reporte de gastos"])
    worksheet.append([f"Usuario: {user.get_full_name() or user.username}"])
    worksheet.append([f"Fecha de generación: {generated_at.strftime('%d/%m/%Y')}"])
    worksheet.append([])
    worksheet.append(headers)

    for expense in expenses:
        worksheet.append([
            expense.title,
            expense.get_category_display(),
            _format_date(expense.date),
            expense.description or "-",
            float(expense.amount),
        ])

    worksheet.append([])
    total_row = worksheet.max_row + 1
    worksheet.cell(row=total_row, column=4, value="Total general")
    worksheet.cell(row=total_row, column=5, value=f"=SUM(E6:E{total_row - 1})")

    header_fill = PatternFill("solid", fgColor="8FA5FF")
    header_font = Font(color="FFFFFF", bold=True)
    total_fill = PatternFill("solid", fgColor="E9EEFF")
    thin_border = Border(
        left=Side(style="thin", color="C9D4FF"),
        right=Side(style="thin", color="C9D4FF"),
        top=Side(style="thin", color="C9D4FF"),
        bottom=Side(style="thin", color="C9D4FF"),
    )

    for cell in worksheet[5]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for row in worksheet.iter_rows(min_row=6, max_row=worksheet.max_row - 1, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border

    for cell in worksheet[total_row]:
        cell.fill = total_fill
        cell.font = Font(bold=True)
        cell.border = thin_border

    worksheet.column_dimensions["A"].width = 34
    worksheet.column_dimensions["B"].width = 18
    worksheet.column_dimensions["C"].width = 14
    worksheet.column_dimensions["D"].width = 45
    worksheet.column_dimensions["E"].width = 14

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
